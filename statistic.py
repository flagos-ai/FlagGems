import json
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def extract_fields(data):
    """从JSON数据中提取所有op的指定字段"""
    rows = []
    try:
        # 获取result对象
        result = data.get("result", {})
        for op_name, op_data in result.items():
            # 提取accuracy部分
            accuracy = op_data.get("accuracy", {})
            if accuracy:
                accuracy_status = accuracy.get("status")
                accuracy_total = accuracy.get("total")
                accuracy_pass_count = accuracy.get("passed")
                accuracy_failed_count = accuracy.get("failed")
                accuracy_skipped_count = accuracy.get("skipped")
                accuracy_errors_count = accuracy.get("errors")
                accuracy_exit_code_count = accuracy.get("exit_code")
            else:
                print("  无 accuracy 数据")
            # 提取performance部分
            performance = op_data.get("performance", {})
            performance_status = performance.get("status")
            performance_data = performance.get("data", {})
            speedup_list = []
            if performance_data is not None:
                if isinstance(performance_data, dict):
                    # 处理bool字段
                    bool_val = performance_data.get("bool")
                    if isinstance(bool_val, dict):
                        bool_val = bool_val.get("speedup")
                        speedup_list.append(bool_val)

                    # 处理int32字段
                    int32_val = performance_data.get("int32")
                    if isinstance(int32_val, dict):
                        int32_val = int32_val.get("speedup")
                        speedup_list.append(int32_val)

                    # 处理fp32字段
                    fp32_val = performance_data.get("fp32")
                    if isinstance(fp32_val, dict):
                        fp32_val = fp32_val.get("speedup")
                        speedup_list.append(fp32_val)

                    # 处理fp16字段
                    fp16_val = performance_data.get("fp16")
                    if isinstance(fp16_val, dict):
                        fp16_val = fp16_val.get("speedup")
                        speedup_list.append(fp16_val)

                    # 处理bf16字段
                    bf16_val = performance_data.get("bf16")
                    if isinstance(bf16_val, dict):
                        bf16_val = bf16_val.get("speedup")
                        speedup_list.append(bf16_val)

                    # 处理int16字段
                    int16_val = performance_data.get("int16")
                    if isinstance(int16_val, dict):
                        int16_val = int16_val.get("speedup")
                        speedup_list.append(int16_val)

                    # 处理cf64字段
                    cf64_val = performance_data.get("cf64")
                    if isinstance(cf64_val, dict):
                        cf64_val = cf64_val.get("speedup")
                        speedup_list.append(cf64_val)

                    # 处理每个算子平均加速比
                    vals = [x for x in speedup_list if x != 0]
                    if not vals:
                        avg_speedup = None
                    else:
                        avg_speedup = sum(vals) / len(vals)
                    # avg_speedup = str(round(float(sum(vals) / len(vals)), 6)) if vals else "0"
                    # 构建行数据
                    row = {
                        "op_name": op_name,
                        "accuracy_status": accuracy_status,
                        "accuracy_total": accuracy_total,
                        "accuracy_pass_count": accuracy_pass_count,
                        "accuracy_failed_count": accuracy_failed_count,
                        "accuracy_skipped_count": accuracy_skipped_count,
                        "accuracy_errors_count": accuracy_errors_count,
                        "accuracy_exit_code_count": accuracy_exit_code_count,
                        "performance_status": performance_status,
                        "bool": bool_val,
                        "int32": int32_val,
                        "fp32": fp32_val,
                        "fp16": fp16_val,
                        "bf16": bf16_val,
                        "int16": int16_val,
                        "cf64": cf64_val,
                        "avg_speedup": avg_speedup
                    }
                    rows.append(row)
                else:
                    row = {
                        "op_name": op_name,
                        "accuracy_status": accuracy_status,
                        "accuracy_total": accuracy_total,
                        "accuracy_pass_count": accuracy_pass_count,
                        "accuracy_failed_count": accuracy_failed_count,
                        "accuracy_skipped_count": accuracy_skipped_count,
                        "accuracy_errors_count": accuracy_errors_count,
                        "accuracy_exit_code_count": accuracy_exit_code_count,
                        "performance_status": performance_status,
                        "bool": None,
                        "int32": None,
                        "fp32": None,
                        "fp16": None,
                        "bf16": None,
                        "int16": None,
                        "cf64": None,
                        "avg_speedup": None
                    }
                    rows.append(row)

    except Exception as e:
        print(f"解析错误: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
    return rows


def json_to_excel(json_file_path, output_excel_path=None):
    """将JSON文件转换为Excel文件"""
    # 读取JSON文件
    print(f"正在读取文件: {json_file_path}")
    try:
        with open(json_file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        print("JSON文件读取成功")
    except FileNotFoundError:
        print(f"错误：文件 {json_file_path} 未找到", file=sys.stderr)
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"错误：JSON解析失败 - {e}", file=sys.stderr)
        sys.exit(1)

    # 提取数据
    print("\n正在解析数据...")
    rows = extract_fields(data)

    if not rows:
        print("警告：未提取到任何数据，请检查JSON结构", file=sys.stderr)
        print("期望的JSON结构：{\"result\": {\"op1\": {...}, \"op2\": {...}, ...}}")
        return

    # 确定输出文件名
    if output_excel_path is None:
        output_excel_path = Path(json_file_path).stem + "_output.xlsx"

    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    # 定义列名
    fieldnames = ["op_name", "accuracy_status", "accuracy_total", "accuracy_pass_count", "accuracy_failed_count", "accuracy_skipped_count", "accuracy_errors_count", "accuracy_exit_code_count", "avg_speedup", "bool", "int32", "fp32", "fp16", "bf16", "int16", "cf64", "performance_status"]
    field_labels = {
        "op_name": "算子名称",
        "accuracy_status": "正确性结果",
        "accuracy_total": "精度测例总数",
        "accuracy_pass_count": "精度测例通过数",
        "accuracy_failed_count": "精度测例失败数",
        "accuracy_skipped_count": "精度测例skip数",
        "accuracy_errors_count": "精度error数",
        "accuracy_exit_code_count": "exit_code数",
        "avg_speedup": "性能加速比",
        "bool": "bool",
        "int32": "int32",
        "fp32": "fp32",
        "fp16": "fp16",
        "bf16": "bf16",
        "int16": "int16",
        "cf64": "cf64",
        "performance_status": "性能结果"
    }

    # 设置样式
    # 表头样式
    header_font = Font(bold=True, size=11, color="2E2E2E")
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 数据单元格样式
    data_alignment = Alignment(horizontal="center", vertical="center")

    # 写入表头
    print("\n正在写入Excel...")
    for col_idx, field in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_idx, value=field_labels.get(field, field))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 写入数据行
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, field in enumerate(fieldnames, 1):
            value = row_data.get(field)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_alignment

    # 设置行高
    ws.row_dimensions[1].height = 20
    for row_idx in range(2, len(rows) + 2):
        ws.row_dimensions[row_idx].height = 18

    # 保存Excel文件
    try:
        wb.save(output_excel_path)
        print(f"\n ✅ 转换成功！")
        print(f" 📁 输出文件：{output_excel_path}")
        print(f" 📊 共处理 {len(rows)} 个操作，{len(fieldnames)} 个字段")

        # 统计信息
        pass_count = sum(1 for row in rows if row.get('status') == 'pass')
        fail_count = sum(1 for row in rows if row.get('status') == 'fail')
        print(f" 📈 统计：通过 {pass_count} 个，失败 {fail_count} 个")
        print(f" 💾 文件大小：{Path(output_excel_path).stat().st_size / 1024:.2f} KB")

    except Exception as e:
        print(f"错误：无法保存Excel文件 - {e}", file=sys.stderr)
        sys.exit(1)


def display_json_structure(json_file_path):
    """显示JSON文件结构（用于调试）"""
    try:
        with open(json_file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        print("\n 📋 JSON文件结构预览：")
        print("-" * 50)

        # 显示result下的所有op
        result = data.get("result", {})
        if result:
            op_count = len(result.keys())
            print(f"✓ result 字段存在，包含 {op_count} 个操作:")
            for op_name, op_data in list(result.items())[:5]:  # 只显示前5个
                acc = op_data.get("acc", {})
                print(f"  - {op_name}: status={acc.get('status')}, total={acc.get('total')}, pass={acc.get('pass')}")

            if op_count > 5:
                print(f"  ... 还有 {op_count - 5} 个操作未显示")
        else:
            print(" ⚠️  未找到 result 字段或 result 为空")
        print("-" * 50)

    except Exception as e:
        print(f"无法预览文件结构：{e}")


def main():
    input_file = "logs_results_20260723_2307/summary.json"
    output_file = "summary_xtriton.xlsx"
    json_to_excel(input_file, output_file)


if __name__ == "__main__":
    main()
