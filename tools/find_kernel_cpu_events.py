#!/usr/bin/env python3
"""
从 PyTorch Profiler 生成的 rank-0.json 文件中找到所有 kernel 及其对应的 CPU event。

用法:
    python find_kernel_cpu_events.py <profile_json_path> [--output <output_file>] [--filter <kernel_name>]

示例:
    python find_kernel_cpu_events.py rank-0.json
    python find_kernel_cpu_events.py rank-0.json --filter matmul
    python find_kernel_cpu_events.py rank-0.json --output kernel_mapping.json
    python find_kernel_cpu_events.py rank-0.json --excel kernel_mapping.xlsx
"""

import argparse
import json
from collections import defaultdict
from typing import Any, Dict, List, Optional

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def load_profile(profile_path: str) -> Dict[str, Any]:
    """加载 profile JSON 文件"""
    with open(profile_path, "r") as f:
        return json.load(f)


def build_cpu_event_index(events: List[Dict]) -> Dict[int, List[Dict]]:
    """构建 External id -> CPU event 的索引"""
    cpu_events = defaultdict(list)
    for event in events:
        cat = event.get("cat", "")
        if cat in ("cpu_op", "user_annotation", "python_function", "ac2g"):
            ext_id = event.get("args", {}).get("External id")
            if ext_id is not None:
                cpu_events[ext_id].append(event)
    return cpu_events


def find_parent_cpu_events(
    target_ext_id: int, cpu_events: Dict[int, List[Dict]], all_events: List[Dict]
) -> List[Dict]:
    """找到给定 External id 的父级 CPU 事件链"""
    target_events = cpu_events.get(target_ext_id, [])
    if not target_events:
        return []

    target_event = target_events[0]
    target_ts = target_event.get("ts", 0)
    target_dur = target_event.get("dur", 0)
    target_tid = target_event.get("tid")

    parents = []
    for event in all_events:
        cat = event.get("cat", "")
        if cat not in ("cpu_op", "user_annotation", "python_function"):
            continue
        if event.get("tid") != target_tid:
            continue

        event_ts = event.get("ts", 0)
        event_dur = event.get("dur", 0)
        event_ext_id = event.get("args", {}).get("External id")

        if event_ext_id == target_ext_id:
            continue
        if event_ts <= target_ts and (event_ts + event_dur) >= (target_ts + target_dur):
            parents.append(event)

    parents.sort(key=lambda e: -(e.get("dur", 0)))
    return parents


def extract_kernel_info(event: Dict) -> Dict:
    """从 kernel 事件中提取关键信息"""
    args = event.get("args", {})
    return {
        "name": event.get("name", ""),
        "timestamp": event.get("ts"),
        "duration_us": event.get("dur"),
        "external_id": args.get("External id"),
        "correlation": args.get("correlation"),
        "device": args.get("device"),
        "stream": args.get("stream"),
        "grid": args.get("grid"),
        "block": args.get("block"),
        "registers_per_thread": args.get("registers per thread"),
        "shared_memory": args.get("shared memory"),
        "occupancy": args.get("est. achieved occupancy %"),
    }


def extract_cpu_event_info(event: Dict) -> Dict:
    """从 CPU 事件中提取关键信息"""
    args = event.get("args", {})
    info = {
        "name": event.get("name", ""),
        "category": event.get("cat", ""),
        "timestamp": event.get("ts"),
        "duration_us": event.get("dur"),
        "external_id": args.get("External id"),
        "thread_id": event.get("tid"),
        "process_id": event.get("pid"),
    }
    if "Call stack" in args:
        info["call_stack"] = args["Call stack"]
    if "Python call stack" in args:
        info["python_call_stack"] = args["Python call stack"]
    if "Input Dims" in args:
        info["input_dims"] = args["Input Dims"]
    if "Input type" in args:
        info["input_type"] = args["Input type"]
    return info


def find_kernel_cpu_mapping(
    profile_path: str,
    kernel_filter: Optional[str] = None,
    include_parents: bool = False,
) -> List[Dict]:
    """找到所有 kernel 和对应的 CPU event 映射"""
    data = load_profile(profile_path)
    events = data.get("traceEvents", [])
    cpu_events = build_cpu_event_index(events)

    results = []
    for event in events:
        if event.get("cat", "") != "kernel":
            continue

        kernel_name = event.get("name", "")
        if kernel_filter and kernel_filter.lower() not in kernel_name.lower():
            continue

        kernel_info = extract_kernel_info(event)
        ext_id = kernel_info["external_id"]
        cpu_event_list = cpu_events.get(ext_id, [])
        cpu_events_info = [extract_cpu_event_info(e) for e in cpu_event_list]

        result = {"kernel": kernel_info, "cpu_events": cpu_events_info}

        if include_parents and cpu_event_list:
            parents = find_parent_cpu_events(ext_id, cpu_events, events)
            result["parent_cpu_events"] = [
                extract_cpu_event_info(e) for e in parents[:5]
            ]

        results.append(result)

    return results


def print_summary(results: List[Dict]):
    """打印摘要信息"""
    print(f"\n{'=' * 80}")
    print(f"找到 {len(results)} 个 kernel 事件")
    print(f"{'=' * 80}\n")

    kernel_stats = defaultdict(lambda: {"count": 0, "total_dur": 0})
    for r in results:
        name = r["kernel"]["name"]
        simple_name = name.split("<")[0] if "<" in name else name
        simple_name = simple_name.split("(")[0] if "(" in simple_name else simple_name
        kernel_stats[simple_name]["count"] += 1
        kernel_stats[simple_name]["total_dur"] += r["kernel"].get("duration_us", 0) or 0

    print("Kernel 统计 (按总耗时排序):")
    print("-" * 80)
    sorted_stats = sorted(kernel_stats.items(), key=lambda x: -x[1]["total_dur"])
    for name, stats in sorted_stats[:20]:
        display_name = name[:55] + "..." if len(name) > 55 else name
        print(
            f"  {display_name:<58} | 次数: {stats['count']:>5} | 总耗时: {stats['total_dur']:>10.2f} us"
        )


def print_detailed(results: List[Dict], limit: int = 50):
    """打印详细的 kernel 到 CPU event 映射"""
    print(f"\n{'=' * 80}")
    print("详细 Kernel -> CPU Event 映射")
    print(f"{'=' * 80}\n")

    for i, r in enumerate(results[:limit]):
        kernel = r["kernel"]
        cpu_events = r["cpu_events"]

        kernel_short = (
            kernel["name"][:80] + "..." if len(kernel["name"]) > 80 else kernel["name"]
        )
        print(f"\n[{i + 1}] Kernel: {kernel_short}")
        print(f"    External ID: {kernel['external_id']}")
        print(f"    Duration: {kernel['duration_us']:.2f} us")
        print(f"    Grid: {kernel['grid']}, Block: {kernel['block']}")

        if cpu_events:
            print("    对应的 CPU 事件:")
            for cpu_e in cpu_events:
                print(
                    f"      - {cpu_e['name']} (cat={cpu_e['category']}, dur={cpu_e.get('duration_us', 'N/A')} us)"
                )
                if "call_stack" in cpu_e:
                    stack_lines = cpu_e["call_stack"].split("\n")[:3]
                    for line in stack_lines:
                        print(f"        {line}")
                if "python_call_stack" in cpu_e:
                    stack_lines = cpu_e["python_call_stack"].split("\n")[:3]
                    for line in stack_lines:
                        print(f"        [py] {line}")
        else:
            print("    对应的 CPU 事件: 无")

        if "parent_cpu_events" in r and r["parent_cpu_events"]:
            print("    父级 CPU 事件链:")
            for parent in r["parent_cpu_events"][:3]:
                print(f"      <- {parent['name']}")


def save_to_excel(results: List[Dict], excel_path: str):
    """将结果保存到 Excel 文件"""
    if not HAS_OPENPYXL:
        print("错误: 需要安装 openpyxl 库来导出 Excel 文件")
        print("请运行: pip install openpyxl")
        return False

    wb = openpyxl.Workbook()

    # === Sheet 1: 详细映射 ===
    ws_detail = wb.active
    ws_detail.title = "Kernel-CPU映射"

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 详细表头
    detail_headers = [
        "序号",
        "Kernel名称",
        "Kernel耗时(us)",
        "External ID",
        "Device",
        "Stream",
        "Grid",
        "Block",
        "Registers/Thread",
        "Shared Memory",
        "Occupancy(%)",
        "CPU Event名称",
        "CPU Event类别",
        "CPU Event耗时(us)",
        "调用栈",
    ]

    for col, header in enumerate(detail_headers, 1):
        cell = ws_detail.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 填充数据
    for idx, r in enumerate(results, 1):
        kernel = r["kernel"]
        cpu_events = r.get("cpu_events", [])

        # CPU event 信息
        cpu_name = ""
        cpu_cat = ""
        cpu_dur = ""
        call_stack = ""
        if cpu_events:
            cpu_e = cpu_events[0]
            cpu_name = cpu_e.get("name", "")
            cpu_cat = cpu_e.get("category", "")
            cpu_dur = cpu_e.get("duration_us", "")
            if "call_stack" in cpu_e:
                call_stack = cpu_e["call_stack"][:500]  # 限制长度
            elif "python_call_stack" in cpu_e:
                call_stack = cpu_e["python_call_stack"][:500]

        row_data = [
            idx,
            kernel.get("name", ""),
            kernel.get("duration_us", ""),
            kernel.get("external_id", ""),
            kernel.get("device", ""),
            kernel.get("stream", ""),
            str(kernel.get("grid", "")),
            str(kernel.get("block", "")),
            kernel.get("registers_per_thread", ""),
            kernel.get("shared_memory", ""),
            kernel.get("occupancy", ""),
            cpu_name,
            cpu_cat,
            cpu_dur,
            call_stack,
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws_detail.cell(row=idx + 1, column=col, value=value)
            cell.border = thin_border
            if col == 15:  # 调用栈列
                cell.alignment = Alignment(wrap_text=True)

    # 设置列宽
    detail_col_widths = [8, 60, 15, 12, 8, 8, 20, 20, 15, 15, 12, 30, 12, 18, 50]
    for col, width in enumerate(detail_col_widths, 1):
        ws_detail.column_dimensions[get_column_letter(col)].width = width

    # === Sheet 2: 统计摘要 ===
    ws_summary = wb.create_sheet(title="统计摘要")

    # 统计数据
    kernel_stats = defaultdict(lambda: {"count": 0, "total_dur": 0, "cpu_ops": set()})
    for r in results:
        name = r["kernel"]["name"]
        simple_name = name.split("<")[0] if "<" in name else name
        simple_name = simple_name.split("(")[0] if "(" in simple_name else simple_name
        kernel_stats[simple_name]["count"] += 1
        kernel_stats[simple_name]["total_dur"] += r["kernel"].get("duration_us", 0) or 0
        for cpu_e in r.get("cpu_events", []):
            kernel_stats[simple_name]["cpu_ops"].add(cpu_e.get("name", ""))

    # 摘要表头
    summary_headers = ["Kernel类型", "调用次数", "总耗时(us)", "平均耗时(us)", "关联的CPU Op"]
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 按总耗时排序
    sorted_stats = sorted(kernel_stats.items(), key=lambda x: -x[1]["total_dur"])
    for idx, (name, stats) in enumerate(sorted_stats, 2):
        avg_dur = stats["total_dur"] / stats["count"] if stats["count"] > 0 else 0
        cpu_ops_str = ", ".join(sorted(stats["cpu_ops"]))

        row_data = [name, stats["count"], stats["total_dur"], avg_dur, cpu_ops_str]
        for col, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=idx, column=col, value=value)
            cell.border = thin_border

    # 设置摘要列宽
    summary_col_widths = [60, 12, 15, 15, 40]
    for col, width in enumerate(summary_col_widths, 1):
        ws_summary.column_dimensions[get_column_letter(col)].width = width

    # 冻结首行
    ws_detail.freeze_panes = "A2"
    ws_summary.freeze_panes = "A2"

    wb.save(excel_path)
    return True


def main():
    parser = argparse.ArgumentParser(
        description="从 profile JSON 中提取 kernel 和 CPU event 的映射关系"
    )
    parser.add_argument("profile_path", help="profile JSON 文件路径")
    parser.add_argument("--output", "-o", help="输出 JSON 文件路径")
    parser.add_argument("--excel", "-e", help="输出 Excel 文件路径")
    parser.add_argument("--filter", "-f", help="按 kernel 名称过滤")
    parser.add_argument("--parents", "-p", action="store_true", help="包含父级 CPU 事件")
    parser.add_argument("--limit", "-l", type=int, default=50, help="详细输出的最大条数")
    parser.add_argument("--summary-only", "-s", action="store_true", help="只输出摘要")

    args = parser.parse_args()

    print(f"加载 profile 文件: {args.profile_path}")
    results = find_kernel_cpu_mapping(args.profile_path, args.filter, args.parents)

    print_summary(results)

    if not args.summary_only:
        print_detailed(results, args.limit)

    if args.output:
        with open(args.output, "w") as f:
            json.dump(results, f, indent=2, ensure_ascii=False)
        print(f"\n结果已保存到: {args.output}")

    if args.excel:
        if save_to_excel(results, args.excel):
            print(f"\nExcel 文件已保存到: {args.excel}")


if __name__ == "__main__":
    main()
